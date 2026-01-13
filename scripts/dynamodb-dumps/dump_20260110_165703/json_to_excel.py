#!/usr/bin/env python3
"""
Convert BudgetBuddy-Transactions.json to Excel file with multiple sheets
"""

import json
import sys
from collections import defaultdict
from decimal import Decimal
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Please install it with: pip3 install openpyxl")
    sys.exit(1)


def extract_dynamodb_value(item, key):
    """Extract value from DynamoDB format"""
    if key not in item:
        return None
    value = item[key]
    if 'S' in value:
        return value['S']
    elif 'N' in value:
        return value['N']
    elif 'BOOL' in value:
        return value['BOOL']
    return None


def format_amount(amount_str):
    """Format amount for display"""
    if not amount_str:
        return None
    try:
        amount = Decimal(amount_str)
        return float(amount)
    except:
        return amount_str


def create_excel_from_transactions(json_file, output_file):
    """Create Excel workbook from transactions JSON"""
    
    print(f"Reading transactions from: {json_file}")
    with open(json_file, 'r') as f:
        data = json.load(f)
    
    items = data.get('Items', [])
    print(f"Total transactions in file: {len(items)}")
    
    # Filter to expense transactions for the user
    user_id = 'c8920e3b-7293-4674-a5f3-fa4c0c6f115a'
    transactions = []
    for item in items:
        uid = extract_dynamodb_value(item, 'userId')
        if uid != user_id:
            continue
        tx_type = extract_dynamodb_value(item, 'transactionType')
        if tx_type != 'EXPENSE':
            continue
        transactions.append(item)
    
    print(f"Expense transactions for user: {len(transactions)}")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Sheet 1: All Transactions
    ws_all = wb.create_sheet("All Transactions")
    headers_all = [
        "Transaction ID", "Date", "Merchant Name", "Description", 
        "Amount", "Category Primary", "Category Detailed",
        "Account ID", "Transaction Type", "Created At",
        "Review Notes", "Is Subscription?", "Is Recurring?"
    ]
    
    ws_all.append(headers_all)
    for cell in ws_all[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    for tx in sorted(transactions, key=lambda x: extract_dynamodb_value(x, 'transactionDate') or '', reverse=True):
        row = [
            extract_dynamodb_value(tx, 'transactionId') or '',
            extract_dynamodb_value(tx, 'transactionDate') or '',
            extract_dynamodb_value(tx, 'merchantName') or '',
            extract_dynamodb_value(tx, 'description') or '',
            format_amount(extract_dynamodb_value(tx, 'amount')),
            extract_dynamodb_value(tx, 'categoryPrimary') or '',
            extract_dynamodb_value(tx, 'categoryDetailed') or '',
            extract_dynamodb_value(tx, 'accountId') or '',
            extract_dynamodb_value(tx, 'transactionType') or '',
            extract_dynamodb_value(tx, 'createdAt') or '',
            '',  # Review Notes - for annotation
            '',  # Is Subscription? - for annotation
            ''   # Is Recurring? - for annotation
        ]
        ws_all.append(row)
        # Add border to data rows
        for cell in ws_all[ws_all.max_row]:
            cell.border = border
    
    # Auto-adjust column widths
    for col in range(1, len(headers_all) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws_all[column_letter]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_all.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Grouped by Merchant (for subscription detection review)
    ws_merchant = wb.create_sheet("Grouped by Merchant")
    headers_merchant = [
        "Merchant Name", "Transaction Count", "Total Amount", 
        "Average Amount", "Min Amount", "Max Amount",
        "Date Range", "Category Primary", "Category Detailed",
        "Sample Transaction IDs", "Is Subscription?", "Is Recurring?", "Review Notes"
    ]
    
    ws_merchant.append(headers_merchant)
    for cell in ws_merchant[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Group transactions by merchant
    by_merchant = defaultdict(list)
    for tx in transactions:
        merchant = extract_dynamodb_value(tx, 'merchantName') or extract_dynamodb_value(tx, 'description') or 'UNKNOWN'
        by_merchant[merchant].append(tx)
    
    # Sort by transaction count (descending)
    sorted_merchants = sorted(by_merchant.items(), key=lambda x: len(x[1]), reverse=True)
    
    for merchant, txs in sorted_merchants:
        if len(txs) < 2:  # Skip single transactions
            continue
        
        amounts = [abs(Decimal(extract_dynamodb_value(tx, 'amount') or '0')) for tx in txs]
        total_amount = sum(amounts)
        avg_amount = total_amount / len(amounts)
        min_amount = min(amounts)
        max_amount = max(amounts)
        
        dates = sorted([extract_dynamodb_value(tx, 'transactionDate') for tx in txs if extract_dynamodb_value(tx, 'transactionDate')])
        date_range = f"{dates[0]} to {dates[-1]}" if dates else ''
        
        # Get most common categories
        categories_primary = [extract_dynamodb_value(tx, 'categoryPrimary') for tx in txs if extract_dynamodb_value(tx, 'categoryPrimary')]
        categories_detailed = [extract_dynamodb_value(tx, 'categoryDetailed') for tx in txs if extract_dynamodb_value(tx, 'categoryDetailed')]
        category_primary = max(set(categories_primary), key=categories_primary.count) if categories_primary else ''
        category_detailed = max(set(categories_detailed), key=categories_detailed.count) if categories_detailed else ''
        
        # Sample transaction IDs (first 3)
        sample_ids = ', '.join([extract_dynamodb_value(tx, 'transactionId') or '' for tx in txs[:3]])
        if len(txs) > 3:
            sample_ids += f' ... ({len(txs)-3} more)'
        
        row = [
            merchant[:100],  # Truncate long merchant names
            len(txs),
            float(total_amount),
            float(avg_amount),
            float(min_amount),
            float(max_amount),
            date_range,
            category_primary,
            category_detailed,
            sample_ids,
            '',  # Is Subscription? - for annotation
            '',  # Is Recurring? - for annotation
            ''   # Review Notes - for annotation
        ]
        ws_merchant.append(row)
        # Add border to data rows
        for cell in ws_merchant[ws_merchant.max_row]:
            cell.border = border
    
    # Auto-adjust column widths
    for col in range(1, len(headers_merchant) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws_merchant[column_letter]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_merchant.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 3: Potential Subscriptions (from evaluated results if available)
    try:
        with open(json_file.replace('.json', '_evaluated.json'), 'r') as f:
            evaluated_data = json.load(f)
        
        ws_evaluated = wb.create_sheet("Evaluated Results")
        headers_evaluated = [
            "Merchant Name", "Amount", "Frequency", "Subscription Type",
            "Subscription Category", "Transaction Count", "Date Range",
            "Transaction IDs", "Correct Category?", "Review Notes"
        ]
        
        ws_evaluated.append(headers_evaluated)
        for cell in ws_evaluated[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Add all evaluated subscriptions/recurring
        all_evaluated = evaluated_data.get('subscriptions', []) + evaluated_data.get('recurring', [])
        for item in sorted(all_evaluated, key=lambda x: abs(Decimal(x.get('amount', '0'))), reverse=True):
            dates = item.get('dates', [])
            date_range = f"{dates[0]} to {dates[-1]}" if dates else ''
            
            # Get sample transaction IDs
            txs = item.get('transactions', [])
            sample_ids = ', '.join([extract_dynamodb_value(tx, 'transactionId') or '' for tx in txs[:3]])
            if len(txs) > 3:
                sample_ids += f' ... ({len(txs)-3} more)'
            
            row = [
                item.get('merchant', ''),
                item.get('amount', ''),
                item.get('frequency', ''),
                item.get('subscription_type', ''),
                item.get('subscription_category', ''),
                item.get('transaction_count', 0),
                date_range,
                sample_ids,
                '',  # Correct Category? - for annotation
                ''   # Review Notes - for annotation
            ]
            ws_evaluated.append(row)
            # Add border to data rows
            for cell in ws_evaluated[ws_evaluated.max_row]:
                cell.border = border
        
        # Auto-adjust column widths
        for col in range(1, len(headers_evaluated) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws_evaluated[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_evaluated.column_dimensions[column_letter].width = adjusted_width
    except FileNotFoundError:
        print("Note: Evaluated results file not found, skipping that sheet")
    
    # Save workbook
    wb.save(output_file)
    print(f"\nExcel file created: {output_file}")
    print(f"Sheets created:")
    for sheet in wb.sheetnames:
        row_count = wb[sheet].max_row - 1  # Exclude header
        print(f"  - {sheet}: {row_count} rows")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        json_file = 'BudgetBuddy-Transactions.json'
        output_file = 'BudgetBuddy-Transactions.xlsx'
    else:
        json_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else json_file.replace('.json', '.xlsx')
    
    create_excel_from_transactions(json_file, output_file)
